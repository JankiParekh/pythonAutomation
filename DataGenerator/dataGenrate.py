import openpyxl

def dataGenerator():
    wk = openpyxl.load_workbook("C:/Janki/Python/TestData.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    finalList = []
    intList = []
    for i in range(1, r+1):
        intList = []
        fname = sh.cell(i,1)
        lname = sh.cell(i,2)
        intList.insert(0, fname.value)
        intList.insert(1, lname.value)
        finalList.insert(i-1,intList)
    print(finalList)
    return finalList
